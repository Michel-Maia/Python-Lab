import openpyxl
import pyperclip
import pyautogui
from time import sleep   # sleep para delay na pg (não ir para outra pg sem carregar)


# Entrar na planilha
workbook = openpyxl.load_workbook('produtos_ficticios.xlsx')
sheet_produtos = workbook['Produtos'] # Produtos é o nome da pg excel onde as informações seram guardadas

# Copiar informações de um campo e colar no seu campo correspondente

#pega a informação salva em uma variavel para salvar no sistema
for linha in sheet_produtos.iter_rows(min_row=2):  # iter_rows passar em cada linha (min_row) linha minima
    # Nome do Produto
    nome_produto = linha[0].value # valor da linha
    #print(linha[0].value)

    # pyperclip usada para ao copiar e colar não apresente erros de caracteres especiais e acentos
    pyperclip.copy(nome_produto) 
   # pyautogui para movimentar o mouse e colar a informação 
    pyautogui.click(328,284,duration=1) #duration é o tempo pra chegar até o local
    pyautogui.hotkey('ctrl','v') #hotkey atalho

    # Descrição
    descricao = linha[1].value
    pyperclip.copy(descricao)
    pyautogui.click(351,422,duration=1)
    pyautogui.hotkey('ctrl','v')
    
    # Categoria
    categoria = linha[2].value
    pyperclip.copy(categoria)
    pyautogui.click(320,533,duration=1)
    pyautogui.hotkey('ctrl','v')

    # Codigo Produto
    codigo_produto = linha[3].value
    pyperclip.copy(codigo_produto)
    pyautogui.click(275,667,duration=1)
    pyautogui.hotkey('ctrl','v')

    # Peso
    peso = linha[4].value
    pyperclip.copy(peso)
    pyautogui.click(240,765,duration=1)
    pyautogui.hotkey('ctrl','v')

    # Dimensões
    dimensoes = linha[5].value
    pyperclip.copy(dimensoes)
    pyautogui.click(237,882,duration=1)
    pyautogui.hotkey('ctrl','v')

    # Botão próximo
    pyautogui.click(198,949,duration=1)
    sleep(3)

    # Preço
    preco = linha[6].value
    pyperclip.copy(preco)
    pyautogui.click(394,317,duration=1)
    pyautogui.hotkey('ctrl','v')

    # Quantidade em estoque
    quantidade_em_estoque = linha[7].value
    pyperclip.copy(quantidade_em_estoque)
    pyautogui.click(338,420,duration=1)
    pyautogui.hotkey('ctrl','v')

    # Data de validade
    data_de_validade = linha[8].value
    pyperclip.copy(data_de_validade)
    pyautogui.click(314,533,duration=1)
    pyautogui.hotkey('ctrl','v')
    
    # Cor
    cor = linha[9].value
    pyperclip.copy(cor)
    pyautogui.click(292,640,duration=1)
    pyautogui.hotkey('ctrl','v')

    # Tamanho  
    # esse campo é um dropdonw 
    tamanho = linha[10].value
    pyautogui.click(259,741,duration=1)  # clica no campo para abrir o drop
    if tamanho == 'Pequeno': # se for pequeno clica em posição
        pyautogui.click(247,783,duration=1)
    elif tamanho == 'Médio': # se for médio clica em posição
        pyautogui.click(238,818,duration=1)
    else:
        pyautogui.click(245,828,duration=1)

    # material    
    material = linha[11].value
    pyperclip.copy(material)
    pyautogui.click(229,836,duration=1)
    pyautogui.hotkey('ctrl','v')
   
    # Botão próximo
    pyautogui.click(183,934,duration=1)

    fabricante = linha[12].value
    pyperclip.copy(fabricante)
    pyautogui.click(166,332,duration=1)
    pyautogui.hotkey('ctrl','v')

    pais_origem = linha[13].value
    pyperclip.copy(pais_origem)
    pyautogui.click(163,443,duration=1)
    pyautogui.hotkey('ctrl','v')

    observacoes = linha[14].value
    pyperclip.copy(observacoes)
    pyautogui.click(164,549,duration=1)
    pyautogui.hotkey('ctrl','v')

    codigo_de_barras = linha[15].value
    pyperclip.copy(codigo_de_barras)
    pyautogui.click(181,711,duration=1)
    pyautogui.hotkey('ctrl','v')

    localizacao_armazem = linha[16].value
    pyperclip.copy(localizacao_armazem)
    pyautogui.click(183,814,duration=1)
    pyautogui.hotkey('ctrl','v')

    # Botão concluir
    pyautogui.click(195,902,duration=1)
    # Botão confirmar inclusão
    pyautogui.click(301,896,duration=1)
    # Botão confirmação 2
    pyautogui.click(1175,230,duration=1)
    # iniciar cadastro novamente
    pyautogui.click(1011,609,duration=1)

# Repetir o passo até finalizar a página
# clicar em próxima
# Repetir os passos e ir para a próxima página (pg 2)
# Repetir os passos e finalizar o cadastro daquele produto e clicar em concluir
# Clicar em ok, para finalizar o processo
# Clicar em ok na msg de confirmação de salvamento no banco de dados
# Clicar em add mais um e repetir o processo até finalizar a planilha
    
    